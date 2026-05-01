import csv
import io

import pytest
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from src.models.user import UserRole
from tests.test_auth import create_test_user, get_auth_header


async def _create_subject_with_lessons(
    client: AsyncClient,
    headers: dict[str, str],
    subject_code: str = "EXP1",
) -> tuple[int, int, int]:
    """Create semester + subject + schedule entry (generates 13 lessons).

    Returns (semester_id, subject_id, schedule_entry_id).
    """
    sem_resp = await client.post(
        "/semesters",
        json={
            "name": "Export Sem",
            "start_date": "2026-02-16",
            "end_date": "2026-05-16",
            "total_weeks": 13,
        },
        headers=headers,
    )
    semester_id = sem_resp.json()["id"]

    subj_resp = await client.post(
        "/subjects",
        json={"name": "ExportSubj", "code": subject_code, "color": "#FF5733"},
        headers=headers,
    )
    subject_id = subj_resp.json()["id"]

    sched_resp = await client.post(
        f"/semesters/{semester_id}/schedule",
        json={
            "subject_id": subject_id,
            "day_of_week": 1,
            "start_time": "09:00",
            "end_time": "10:40",
            "room": "B213",
            "lesson_type": "cvicenie",
        },
        headers=headers,
    )
    entry_id = sched_resp.json()["id"]

    return semester_id, subject_id, entry_id


@pytest.mark.asyncio
async def test_export_students_csv(
    test_client: AsyncClient, db_session: AsyncSession
) -> None:
    await create_test_user(
        db_session, "admin@exp1.sk", "pass", role=UserRole.admin,
        first_name="Admin", last_name="Exp1",
    )
    headers = await get_auth_header(test_client, "admin@exp1.sk", "pass")
    _, subject_id, _ = await _create_subject_with_lessons(
        test_client, headers, "EX01"
    )

    # Enroll 2 students — one with diacritics
    await test_client.post(
        f"/subjects/{subject_id}/students",
        json={
            "isic_identifier": "110000001",
            "first_name": "\u013dubom\u00edr",
            "last_name": "Kov\u00e1\u010d",
        },
        headers=headers,
    )
    await test_client.post(
        f"/subjects/{subject_id}/students",
        json={
            "isic_identifier": "110000002",
            "first_name": "Maria",
            "last_name": "Banicka",
        },
        headers=headers,
    )

    resp = await test_client.get(
        f"/subjects/{subject_id}/export/students?format=csv",
        headers=headers,
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "text/csv; charset=utf-8"
    assert "attachment; filename=" in resp.headers["content-disposition"]
    assert resp.headers["content-disposition"].endswith(".csv")

    body = resp.content.decode("utf-8-sig")
    lines = body.strip().split("\r\n")
    assert lines[0] == "ISIC,Meno,Priezvisko"
    assert len(lines) == 3  # header + 2 data rows

    # Verify diacritics preserved (sorted by last_name: Banicka before Kováč)
    assert "110000002,Maria,Banicka" in lines[1]
    assert "110000001,\u013dubom\u00edr,Kov\u00e1\u010d" in lines[2]


@pytest.mark.asyncio
async def test_export_students_csv_for_one_time_subject(
    test_client: AsyncClient, db_session: AsyncSession
) -> None:
    await create_test_user(
        db_session, "admin@exp-one-time.sk", "pass", role=UserRole.admin,
        first_name="Admin", last_name="ExpOneTime",
    )
    headers = await get_auth_header(test_client, "admin@exp-one-time.sk", "pass")

    sem_resp = await test_client.post(
        "/semesters",
        json={
            "name": "Export One Time Sem",
            "start_date": "2026-02-16",
            "end_date": "2026-05-16",
            "total_weeks": 13,
        },
        headers=headers,
    )
    semester_id = sem_resp.json()["id"]

    subj_resp = await test_client.post(
        "/subjects",
        json={
            "name": "Export One Time",
            "code": "EXOT",
            "color": "#FF5733",
        },
        headers=headers,
    )
    subject_id = subj_resp.json()["id"]

    sched_resp = await test_client.post(
        f"/semesters/{semester_id}/schedule",
        json={
            "subject_id": subject_id,
            "day_of_week": 3,
            "start_time": "11:00",
            "end_time": "12:40",
            "room": "B312",
            "lesson_type": "laboratorium",
            "is_one_time": True,
            "recurrence_interval": 1,
        },
        headers=headers,
    )
    assert sched_resp.status_code == 201

    enroll_resp = await test_client.post(
        f"/subjects/{subject_id}/students",
        json={
            "isic_identifier": "110000099",
            "first_name": "Jednorazova",
            "last_name": "Studentka",
        },
        headers=headers,
    )
    assert enroll_resp.status_code == 201

    resp = await test_client.get(
        f"/subjects/{subject_id}/export/students?format=csv",
        headers=headers,
    )
    assert resp.status_code == 200
    body = resp.content.decode("utf-8-sig")
    assert body.strip().split("\r\n") == [
        "ISIC,Meno,Priezvisko",
        "110000099,Jednorazova,Studentka",
    ]


@pytest.mark.asyncio
async def test_export_students_xlsx(
    test_client: AsyncClient, db_session: AsyncSession
) -> None:
    await create_test_user(
        db_session, "admin@exp2.sk", "pass", role=UserRole.admin,
        first_name="Admin", last_name="Exp2",
    )
    headers = await get_auth_header(test_client, "admin@exp2.sk", "pass")
    _, subject_id, _ = await _create_subject_with_lessons(
        test_client, headers, "EX02"
    )

    await test_client.post(
        f"/subjects/{subject_id}/students",
        json={
            "isic_identifier": "120000001",
            "first_name": "Anna",
            "last_name": "Horvathova",
        },
        headers=headers,
    )
    await test_client.post(
        f"/subjects/{subject_id}/students",
        json={
            "isic_identifier": "120000002",
            "first_name": "Peter",
            "last_name": "Kovac",
        },
        headers=headers,
    )

    resp = await test_client.get(
        f"/subjects/{subject_id}/export/students?format=xlsx",
        headers=headers,
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert resp.headers["content-disposition"].endswith(".xlsx")

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 3  # header + 2 data
    assert rows[0] == ("ISIC", "Meno", "Priezvisko")


@pytest.mark.asyncio
async def test_export_attendance_csv(
    test_client: AsyncClient, db_session: AsyncSession
) -> None:
    await create_test_user(
        db_session, "admin@exp3.sk", "pass", role=UserRole.admin,
        first_name="Admin", last_name="Exp3",
    )
    headers = await get_auth_header(test_client, "admin@exp3.sk", "pass")
    semester_id, subject_id, entry_id = await _create_subject_with_lessons(
        test_client, headers, "EX03"
    )

    # Enroll 2 students
    enroll1 = await test_client.post(
        f"/subjects/{subject_id}/students",
        json={
            "isic_identifier": "130000001",
            "first_name": "Student",
            "last_name": "Alpha",
        },
        headers=headers,
    )
    assert enroll1.status_code == 201

    enroll2 = await test_client.post(
        f"/subjects/{subject_id}/students",
        json={
            "isic_identifier": "130000002",
            "first_name": "Student",
            "last_name": "Beta",
        },
        headers=headers,
    )
    assert enroll2.status_code == 201

    # Change student Alpha's attendance for week 1 to pritomny
    lessons_resp = await test_client.get(
        f"/semesters/{semester_id}/schedule/{entry_id}/lessons",
        headers=headers,
    )
    lesson_id = lessons_resp.json()[0]["id"]

    att_resp = await test_client.get(
        f"/lessons/{lesson_id}/attendance", headers=headers
    )
    students = att_resp.json()["students"]
    alpha_att = next(
        s for s in students if s["isic_identifier"] == "130000001"
    )
    await test_client.patch(
        f"/attendance/{alpha_att['attendance_id']}",
        json={"status": "pritomny"},
        headers=headers,
    )

    resp = await test_client.get(
        f"/subjects/{subject_id}/export/attendance?semester_id={semester_id}&format=csv",
        headers=headers,
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "text/csv; charset=utf-8"

    body = resp.content.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(body)))
    header_cols = rows[0]

    # Header: student id, ISIC, name fields, then 13 lesson columns
    assert header_cols[0] == "Študent ID"
    assert header_cols[1] == "ISIC"
    assert header_cols[2] == "Meno"
    assert header_cols[3] == "Priezvisko"
    assert len(header_cols) == 4 + 13
    assert header_cols[4] == (
        "Týždeň 1 | 2026-02-16 | Cvičenie | 09:00-10:40 B213"
    )

    # 2 student rows
    assert len(rows) == 3

    # Find Alpha row (sorted by last_name: Alpha before Beta)
    alpha_cols = rows[1]
    assert alpha_cols[0] == str(enroll1.json()["isic_id"])
    assert alpha_cols[1] == "130000001"
    assert alpha_cols[4] == "Prítomný"  # T1 changed

    # Beta row - all nepritomny
    beta_cols = rows[2]
    assert beta_cols[0] == str(enroll2.json()["isic_id"])
    assert beta_cols[1] == "130000002"
    assert beta_cols[4] == "Neprítomný"


@pytest.mark.asyncio
async def test_export_attendance_includes_all_subject_lesson_types(
    test_client: AsyncClient, db_session: AsyncSession
) -> None:
    await create_test_user(
        db_session, "admin@exp-types.sk", "pass", role=UserRole.admin,
        first_name="Admin", last_name="ExpTypes",
    )
    headers = await get_auth_header(test_client, "admin@exp-types.sk", "pass")
    semester_id, subject_id, _ = await _create_subject_with_lessons(
        test_client, headers, "EXTY"
    )

    for day, lesson_type, room in [
        (2, "prednaska", "A101"),
        (3, "laboratorium", "L204"),
    ]:
        schedule_resp = await test_client.post(
            f"/semesters/{semester_id}/schedule",
            json={
                "subject_id": subject_id,
                "day_of_week": day,
                "start_time": "11:00",
                "end_time": "12:40",
                "room": room,
                "lesson_type": lesson_type,
            },
            headers=headers,
        )
        assert schedule_resp.status_code == 201

    enroll_resp = await test_client.post(
        f"/subjects/{subject_id}/students",
        json={
            "isic_identifier": "131000001",
            "first_name": "Full",
            "last_name": "Export",
        },
        headers=headers,
    )
    assert enroll_resp.status_code == 201

    resp = await test_client.get(
        f"/subjects/{subject_id}/export/attendance?semester_id={semester_id}&format=csv",
        headers=headers,
    )
    assert resp.status_code == 200

    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8-sig"))))
    header_cols = rows[0]
    assert len(header_cols) == 4 + (13 * 3)
    assert "Týždeň 1 | 2026-02-16 | Cvičenie | 09:00-10:40 B213" in header_cols
    assert "Týždeň 1 | 2026-02-17 | Prednáška | 11:00-12:40 A101" in header_cols
    assert "Týždeň 1 | 2026-02-18 | Laboratórium | 11:00-12:40 L204" in header_cols
    assert rows[1][0] == str(enroll_resp.json()["isic_id"])
    assert rows[1][1] == "131000001"


@pytest.mark.asyncio
async def test_export_diacritics(
    test_client: AsyncClient, db_session: AsyncSession
) -> None:
    await create_test_user(
        db_session, "admin@exp4.sk", "pass", role=UserRole.admin,
        first_name="Admin", last_name="Exp4",
    )
    headers = await get_auth_header(test_client, "admin@exp4.sk", "pass")
    _, subject_id, _ = await _create_subject_with_lessons(
        test_client, headers, "EX04"
    )

    await test_client.post(
        f"/subjects/{subject_id}/students",
        json={
            "isic_identifier": "140000001",
            "first_name": "\u013dubom\u00edr",
            "last_name": "Kov\u00e1\u010d",
        },
        headers=headers,
    )

    resp = await test_client.get(
        f"/subjects/{subject_id}/export/students?format=csv",
        headers=headers,
    )
    assert resp.status_code == 200

    body = resp.content.decode("utf-8-sig")
    assert "\u013dubom\u00edr" in body
    assert "Kov\u00e1\u010d" in body


@pytest.mark.asyncio
async def test_export_empty_subject(
    test_client: AsyncClient, db_session: AsyncSession
) -> None:
    await create_test_user(
        db_session, "admin@exp5.sk", "pass", role=UserRole.admin,
        first_name="Admin", last_name="Exp5",
    )
    headers = await get_auth_header(test_client, "admin@exp5.sk", "pass")
    _, subject_id, _ = await _create_subject_with_lessons(
        test_client, headers, "EX05"
    )

    resp = await test_client.get(
        f"/subjects/{subject_id}/export/students?format=csv",
        headers=headers,
    )
    assert resp.status_code == 200

    body = resp.content.decode("utf-8-sig")
    assert body == "ISIC,Meno,Priezvisko\r\n"


@pytest.mark.asyncio
async def test_export_large_streaming(
    test_client: AsyncClient, db_session: AsyncSession
) -> None:
    await create_test_user(
        db_session, "admin@exp6.sk", "pass", role=UserRole.admin,
        first_name="Admin", last_name="Exp6",
    )
    headers = await get_auth_header(test_client, "admin@exp6.sk", "pass")
    _, subject_id, _ = await _create_subject_with_lessons(
        test_client, headers, "EX06"
    )

    # Enroll 100 students
    for i in range(100):
        resp = await test_client.post(
            f"/subjects/{subject_id}/students",
            json={
                "isic_identifier": f"16{i:07d}",
                "first_name": f"First{i}",
                "last_name": f"Last{i:03d}",
            },
            headers=headers,
        )
        assert resp.status_code == 201

    resp = await test_client.get(
        f"/subjects/{subject_id}/export/students?format=csv",
        headers=headers,
    )
    assert resp.status_code == 200

    body = resp.content.decode("utf-8-sig")
    lines = body.strip().split("\r\n")
    assert len(lines) == 101  # header + 100 data rows


@pytest.mark.asyncio
async def test_export_download_headers(
    test_client: AsyncClient, db_session: AsyncSession
) -> None:
    await create_test_user(
        db_session, "admin@exp7.sk", "pass", role=UserRole.admin,
        first_name="Admin", last_name="Exp7",
    )
    headers = await get_auth_header(test_client, "admin@exp7.sk", "pass")
    _, subject_id, _ = await _create_subject_with_lessons(
        test_client, headers, "EX07"
    )

    resp = await test_client.get(
        f"/subjects/{subject_id}/export/students?format=csv",
        headers=headers,
    )
    assert resp.status_code == 200
    assert resp.headers["content-disposition"].startswith("attachment; filename=")
    assert resp.headers["content-type"] == "text/csv; charset=utf-8"


@pytest.mark.asyncio
async def test_openapi_schema(test_client: AsyncClient) -> None:
    response = await test_client.get("/openapi.json")
    assert response.status_code == 200
    schema = response.json()
    paths = schema["paths"]

    assert "/subjects/{subject_id}/export/students" in paths
    assert "/subjects/{subject_id}/export/attendance" in paths

    all_tags: set[str] = set()
    for path_ops in paths.values():
        for op in path_ops.values():
            if isinstance(op, dict) and "tags" in op:
                all_tags.update(op["tags"])

    assert "export" in all_tags


@pytest.mark.asyncio
async def test_openapi_full_coverage(test_client: AsyncClient) -> None:
    response = await test_client.get("/openapi.json")
    assert response.status_code == 200
    schema = response.json()
    paths = schema["paths"]

    expected_paths = [
        "/health",
        "/auth/login",
        "/auth/me",
        "/auth/register",
        "/scans",
        "/semesters",
        "/subjects",
        "/subjects/{subject_id}/students",
        "/subjects/{subject_id}/students/import",
        "/subjects/{subject_id}/students/{enrollment_id}",
        "/subjects/{subject_id}/export/students",
        "/subjects/{subject_id}/export/attendance",
    ]
    for path in expected_paths:
        assert path in paths, f"Missing path: {path}"

    all_tags: set[str] = set()
    for path_ops in paths.values():
        for op in path_ops.values():
            if isinstance(op, dict) and "tags" in op:
                all_tags.update(op["tags"])

    expected_tags = [
        "scans", "auth", "semesters", "subjects", "schedule",
        "weeks", "students", "lessons", "attendance", "export",
    ]
    for tag in expected_tags:
        assert tag in all_tags, f"Missing tag: {tag}"
